import os
import argparse
import sys
import zipfile
from openpyxl import load_workbook, Workbook
import re
from pathlib import Path, PurePosixPath
from glob import glob
import shutil
import time

DEFECT_NAME_MAP = {
    'Misaligned or Missing Balls or Columns': 'Misaligned/Missing Balls/Columns | Mechanical | Lead/Balls/Columns',
    'No ESD Bags for ESD Sensitive Devices': 'No ESD Bags for ESD Sensitive Devices | Procedural | Packaging/Shipping',
    'Lead Re-attachment': 'Lead Re-attachment | Mechanical | Lead/Balls/Columns',
    'Defect Name': 'Defect Name',
    'Oxidation or Corrosion': 'Oxidation/Corrosion | Environmental | Lead/Balls/Columns',
    'Die Damage or Extraneous Markings': 'Die Damage or Extraneous Markings | Mechanical | Die',
    'Delamination': 'Delamination | Mechanical | Die',
    'Mechanical Interfaces Intermetallic Growth': 'Mechanical Interfaces: Intermetallic Growth | Electrical | Manufacturing | Package',
    'Improper Textures': 'Improper Textures | Mechanical | Package',
    'Pin Marker Filled or Missing': 'Pin Marker Filled/Missing | Mechanical | Package',
    'Surface Impurities': 'Surface Impurities | Electrical | Manufacturing | Material',
    'Fine or Gross Leak (Hermetic)': 'Fine/Gross Leak (Hermetic) | Mechanical | Package',
    'Threshold Variation': 'Threshold Variation | Electrical | Parametric',
    'Package Contamination': 'Package Contamination | Environmental | Package',
    'Invalid OCM or OEM Shipping Labels': 'Invalid OCM/OEM Shipping Labels | Procedural | Packaging/Shipping',
    'Parasitic Transistors': 'Parasitic Transistors | Electrical | Manufacturing | Process',
    'Invalid Lot Code': 'Invalid Lot Code | Procedural | Packaging/Shipping',
    'Tooling Marks': 'Tooling Marks | Mechanical | Lead/Balls/Columns',
    'Package Mold Variations': 'Package Mold Variations | Mechanical | Package',
    'No Moisture Protection for Moisture Sensitive Devices': 'No Moisture Protection for Moisture Sensitive Devices | Procedural | Packaging/Shipping',
    'Incorrect Dimensions or Weight': 'Incorrect Dimensions/Weight | Mechanical | Package',
    'TDDB': 'TDDB | Electrical | Parametric',
    'Markings': 'Markings | Mechanical | Package',
    'Dirty Cavities': 'Dirty Cavities | Mechanical | Package',
    'Surface Passivation and Corrosion': 'Surface Passivation and Corrosion | Electrical | Manufacturing | Package',
    'Contamination': 'Contamination | Environmental | Lead/Balls/Columns',
    'Abnormal Package Conditions': 'Abnormal Package Conditions | Environmental | Package',
    'Resistive Open or Short': 'Resistive Open/Short | Electrical | Parametric',
    'Sanding or Grinding Marks': 'Sanding/Grinding Marks | Mechanical | Package',
    'Incorrect Dimensions': 'Incorrect Dimensions | Mechanical | Lead/Balls/Columns',
    'Resurfacing or Blacktopping': 'Resurfacing/Blacktopping | Mechanical | Package',
    'Poor Connection': 'Poor Connection | Mechanical | Bond Wires',
    'Out-of-Spec Leakage Current': 'Out-of-Spec Leakage Current | Electrical | Parametric',
    'Crystal Imperfection': 'Crystal Imperfection | Electrical | Manufacturing | Material',
    'Multiple Date Codes Within Lot': 'Multiple Date Codes Within Lot | Procedural | Packaging/Shipping',
    'Mechanical Interfaces Fatigue': 'Mechanical Interfaces: Fatigue | Electrical | Manufacturing | Package',
    'Invalid OCM or OEM Packaging': 'Invalid OCM or OEM Packaging | Procedural | Packaging/Shipping',
    'Wrong Die': 'Wrong Die | Mechanical | Die',
    'Color Variations or Fade': 'Color Variations/Fade | Mechanical | Package',
    'Improper Materials (Seals,Epoxies, Dielectrics, etc)': 'Improper Materials (Seals,Epoxies, Dielectrics, etc.) | Electrical | Manufacturing | Material',
    'Invalid Markings': 'Invalid Markings | Procedural | Package',
    'Package Damage': 'Package Damage | Mechanical | Package',
    'Incorrect Temp Profile': 'Incorrect Temp. Profile | Electrical | Parametric',
    'Poor or Inconsistent Die Attach': 'Poor/Inconsistent Die Attach | Mechanical | Die',
    'Poor or Inconsistent Lead Dress or Lead Frame': 'Poor/Inconsistent Lead Dress/Lead Frame | Mechanical | Bond Wires',
    'Oxide Breakdown': 'Oxide Breakdown | Electrical | Manufacturing | Process',
    'Re-Worked Wire Bonds': 'Re-Worked Wire Bonds | Mechanical | Bond Wires',
    'Missing Wires': 'Missing Wires | Mechanical | Bond Wires',
    'Bond fPull Strength': 'Bond Pull Strength | Mechanical | Bond Wires',
    'Color Variations': 'Color Variations | Mechanical | Lead/Balls/Columns',
    'Reworked': 'Reworked | Mechanical | Lead/Balls/Columns',
    'Ghost Markings': 'Ghost Markings | Mechanical | Package',
    'Missing Contact Windows': 'Missing Contact Windows | Electrical | Manufacturing | Process',
    'Distorted Non-uniform Balls or Columns': 'Distorted Non-uniform Balls/Columns | Mechanical | Lead/Balls/Columns',
    'Wrong Materials': 'Wrong Materials | Mechanical | Lead/Balls/Columns',
    'Missing Die': 'Missing Die | Mechanical | Die',
    'Fine Cracks': 'Fine Cracks | Electrical | Manufacturing | Material',
    'Dents, Damages, or Bent': 'Dents, Damages, or Bent | Mechanical | Lead/Balls/Columns',
    'Burned Markings': 'Burned Markings | Mechanical | Package',
    'Part Orientation within Packaging': 'Part Orientation within Packaging | Procedural | Packaging/Shipping',
    'Missing or Forged Paperwork': 'Missing or Forged Paperwork | Procedural | Packaging/Shipping',
    'Extraneous Markings': 'Extraneous Markings | Mechanical | Package',
    'Broken Wires': 'Broken Wires | Mechanical | Bond Wires',
    'Delay Defects': 'Delay Defects | Electrical | Parametric',
    'Gross Cracks': 'Gross Cracks | Mechanical | Die',
    'Improper Die Markings': 'Improper Die Markings | Mechanical | Die',
    'Misaligned Window': 'Misaligned Window | Electrical | Manufacturing | Process'
}

DEFECT_TEMP_MAP = {
    "Dirty Cavities": "Dirty Cavities",
    "Distorted Pins": "Distorted Non-uniform Balls or Columns",
    "Extraneous Markings": "Extraneous Markings",
    "Misaligned Pins": "Misaligned or Missing Balls or Columns",
    'Misaligned': "Misaligned or Missing Balls or Columns",
    "Oxidation Corrosion": "Oxidation or Corrosion",
    "Oxidation or Corrosion": "Oxidation or Corrosion",
    "Color Variations": "Color Variations",
    "Contamination Pin": "Contamination",
    "Contamination": "Contamination",
    "Contamination Pins": "Contamination",
    "Improper Textures": "Improper Textures",
    "Package Mold Variations": "Package Mold Variations",
    "Corrosion": "Surface Passivation and Corrosion",
    "Package Damage": 'Package Damage',
    "Fine or Gross Leak": 'Fine or Gross Leak (Hermetic)',
    'Sanding Grinding Marks': 'Sanding or Grinding Marks',
    'Dents': 'Dents, Damages, or Bent',
    'Markings': 'Markings',
    'Color Variaitions-Fade': 'Color Variations or Fade',
    'Ghost Markings': 'Ghost Markings',
    'Texture Variations': 'Improper Textures',
    'Texture Variation': 'Improper Textures',
    'Tooling Marks': 'Tooling Marks',
    'Burned Markings': 'Burned Markings'
}

DEFECT_SEQUENCE_MAP = {
    'ML7': 'Misaligned or Missing Balls or Columns',
    2: 'No ESD Bags for ESD Sensitive Devices',
    3: 'Lead Re-attachment',
    4: 'Defect Name',
    5: 'Oxidation or Corrosion',
    6: 'Die Damage or Extraneous Markings',
    7: 'Delamination',
    8: 'Mechanical Interfaces Intermetallic Growth',
    9: 'Improper Textures',
    10: 'Pin Marker Filled or Missing',
    11: 'Surface Impurities',
    12: 'Fine or Gross Leak (Hermetic)',
    13: 'Threshold Variation',
    14: 'Package Contamination',
    15: 'Invalid OCM or OEM Shipping Labels',
    16: 'Parasitic Transistors',
    17: 'Invalid Lot Code',
    18: 'Tooling Marks',
    19: 'Package Mold Variations',
    20: 'No Moisture Protection for Moisture Sensitive Devices',
    21: 'Incorrect Dimensions or Weight',
    22: 'TDDB',
    23: 'Markings',
    24: 'Dirty Cavities',
    25: 'Surface Passivation and Corrosion',
    26: 'Contamination',
    27: 'Abnormal Package Conditions',
    28: 'Resistive Open or Short',
    29: 'Sanding or Grinding Marks',
    30: 'Incorrect Dimensions',
    31: 'Resurfacing or Blacktopping',
    32: 'Poor Connection',
    33: 'Out-of-Spec Leakage Current',
    34: 'Crystal Imperfection',
    35: 'Multiple Date Codes Within Lot',
    36: 'Mechanical Interfaces Fatigue',
    37: 'Invalid OCM or OEM Packaging',
    38: 'Wrong Die',
    39: 'Color Variations or Fade',
    40: 'Improper Materials (Seals,Epoxies, Dielectrics, etc)',
    41: 'Invalid Markings',
    42: 'Package Damage',
    43: 'Incorrect Temp Profile',
    44: 'Poor or Inconsistent Die Attach',
    45: 'Poor or Inconsistent Lead Dress or Lead Frame',
    46: 'Oxide Breakdown',
    47: 'Re-Worked Wire Bonds',
    48: 'Missing Wires',
    49: 'Bond fPull Strength',
    50: 'Color Variations',
    51: 'Reworked',
    52: 'Ghost Markings',
    53: 'Missing Contact Windows',
    54: 'Distorted Non-uniform Balls or Columns',
    55: 'Wrong Materials',
    56: 'Missing Die',
    57: 'Fine Cracks',
    58: 'Dents, Damaged, or Bent',
    59: 'Burned Markings',
    60: 'Part Orientation within Packaging',
    61: 'Missing or Forged Paperwork',
    62: 'Extraneous Markings',
    63: 'Broken Wires',
    64: 'Delay Defects',
    65: 'Gross Cracks',
    66: 'Improper Die Markings',
    67: 'Misaligned Window'
}


ACCEPTED_IMAGES = tuple('.jpg .jpe .jpeg .png'.split())
# SAMPLE_SUBMIT_FORM = Path('sampleSubmitForm.xlsx')
ZIP_IMAGE_DIR = Path('images')
CURRENT_DIR = Path(__file__).parent.absolute()
ROW_START = 68
SAMPLE_ID = 1


def create_workbook(base_dir, files_to_zip, zip_path_name, upload_workbook_path, defect_name_map):
    wb = Workbook()
    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    row = ROW_START
    for image_path, arcname, sample_id in files_to_zip:
        ws.cell(row=row, column=1).value = sample_id
        ws.cell(
            row=row, column=2).value = defect_name_map[DEFECT_TEMP_MAP[image_path.parent.name]]
        ws.cell(row=row, column=3).value = '{}'.format(arcname)
        row += 1
    wb.save(upload_workbook_path)
    return upload_workbook_path


def create_zip_archive(base_dir, product_dir, files_to_zip, defect_name_map):
    manufacturer_dir = product_dir.parent.name
    zip_path_name = manufacturer_dir + '_' + product_dir.name + \
        '_' + str(time.time()).replace('.', '_')
    upload_workbook_path = Path(zip_path_name + '.xlsx')
    zip_name = Path(zip_path_name + '.zip')
    xfile = create_workbook(base_dir, files_to_zip, zip_path_name,
                            upload_workbook_path, defect_name_map)
    with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zf:
        for image_path, arcname, sample_id in files_to_zip:
            zf.write(image_path, arcname)
        zf.write(xfile, os.path.basename(xfile))
    shutil.move(zip_name, base_dir.joinpath(
        manufacturer_dir).joinpath(product_dir).joinpath(zip_name))
    Path(xfile).unlink()


def create_image_upload_archive(path, base_dir, zip_image_dir, accepted_images, defect_name_map):
    product_dir = Path()
    total_size = 0.0
    files_to_zip = []
    for file in Path(path).rglob('*'):
        if file.is_file() and (file.suffix in accepted_images) and (file.parent.name in DEFECT_TEMP_MAP.keys()):
            if not file.parent.parent.parent.samefile(product_dir):
                if files_to_zip:
                    create_zip_archive(
                        base_dir, product_dir, files_to_zip, defect_name_map)
                product_dir = file.parent.parent.parent
                total_size = 0.0
                files_to_zip = []
            if file.parent.parent.parent.samefile(product_dir):
                file_size = file.stat().st_size
                if ((file_size / 1048576.0) > 100.0):
                    continue
                if ((total_size + file_size) / 1048576.0) > 100.0:
                    create_zip_archive(
                        base_dir, product_dir, files_to_zip, defect_name_map)
                    files_to_zip.clear()
                    total_size = 0.0
                if ((total_size + file_size) / 1048576.0) < 100.0:
                    total_size += file_size
                    sub = PurePosixPath(file.relative_to(base_dir))
                    sub_name = str(sub).replace(' ', '_').replace("/", '_')
                    file_sample_id_num = -1
                    sample_dir = file.parent.parent.name
                    try:
                        file_sample_id_num = int(sample_dir)
                    except Exception as e:
                        print("Invalid Sample ID for Product ", product_dir.name)
                        pass
                    arcname = PurePosixPath(zip_image_dir).joinpath(sub_name)
                    print("files_to_zip ", file)
                    files_to_zip.append(
                        (file, arcname, file_sample_id_num))

    if files_to_zip:
        create_zip_archive(base_dir, product_dir,
                           files_to_zip, defect_name_map)
        total_size = 0.0
        files_to_zip.clear()


def create_archive(path):
    path = Path(path).resolve()
    BASE_DIR = Path(path).resolve()
    create_image_upload_archive(path, BASE_DIR, ZIP_IMAGE_DIR,
                                ACCEPTED_IMAGES, DEFECT_NAME_MAP)


if __name__ == "__main__":
    print(DEFECT_NAME_MAP.keys())
    # try:
    #     path = input(
    #         "Enter the absolute path of the folder containing the images to be zipped: ")
    #     if not path:
    #         raise Exception(
    #             'Please Enter a valid path containing the images to be zipped!')
    #     create_archive(path)
    # except Exception as e:
    #     print('------------------ERROR--------------------------')
    #     print(e)
    #     print('-------------------------------------------------')
