from counterfeit_ic import app, db, pimage, pspec, archives
from flask import render_template, redirect, session, request, url_for, flash, abort, send_file, jsonify
from openpyxl import load_workbook, Workbook
from user.models import User
from inventory.models import Manufacturer, Product, DefectType, Chip, Sample, Defect
from inventory.models import ProductSchema, DefectTypeSchema
from inventory.forms import CreateComponentsForm, EditProductForm
from sqlalchemy.sql import func
from werkzeug.utils import secure_filename
from user.decorators import contributor_required, approval_required

import pyexcel as excel
import zipfile
import os
import time
import shutil
import glob
import datetime
import pathlib
from io import BytesIO
import logging


@app.route('/view/products/select/', methods=('GET', 'POST'))
@approval_required
def select_view_product():
    manufacturers = Manufacturer.query.all()
    products = Product.query.all()
    return render_template(
        'inventory/select_product_for_report.html',
        manufacturers=manufacturers,
        products=products,
        form_method='product_report'
    )


@app.route('/edit/products/select/', methods=('GET', 'POST'))
@contributor_required
def select_edit_product():
    manufacturers = Manufacturer.query.all()
    products = Product.query.all()
    return render_template(
        'inventory/select_product_for_edit.html',
        manufacturers=manufacturers,
        products=products,
        form_method='display_edit_product'
    )


@app.route('/edit/defects/')
@contributor_required
def select_defect_sample():
    samples = None
    if session.get('is_admin'):
        samples = db.session.query(
            Manufacturer.name.label('manufacturer'),
            Product.name.label('product'),
            User.username.label('username'),
            Chip.id.label('chip_id'),
            Sample.id.label('sample_id'),
            Sample.sample_id.label('sample_sample_id'),
            Sample.creation_date.label('sample_creation_date'),
        ).filter(
            Chip.id == Sample.chip
        ).filter(
            Chip.product == Product.id
        ).filter(
            Chip.manufacturer == Manufacturer.id
        ).filter(
            Chip.user == User.id
        ).all()
    else:
        user = session.get('id')
        samples = db.session.query(
            Manufacturer.name.label('manufacturer'),
            Product.name.label('product'),
            User.username.label('username'),
            Chip.id.label('chip_id'),
            Sample.id.label('sample_id'),
            Sample.sample_id.label('sample_sample_id'),
            Sample.creation_date.label('sample_creation_date'),
        ).filter(
            Chip.id == Sample.chip
        ).filter(
            Chip.user == user
        ).filter(
            Chip.product == Product.id
        ).filter(
            Chip.manufacturer == Manufacturer.id
        ).filter(
            Chip.user == User.id
        ).all()
    return render_template(
        'inventory/edit_defect.html',
        samples=samples,
    )


@app.route('/edit/defects/', methods=('POST',))
@contributor_required
def edit_defect_sample():
    if (request.method == 'POST'):
        sample_id = request.form.get('sample_id')
        temp = app.config['TEMP']
        path = os.path.join(app.config['APP_ROOT'], 'static')
        old_working_directory = os.getcwd()
        os.chdir(path)
        row = app.config['ROW_START']
        file_path = app.config['FILE_PATH']
        sample_upload_file = app.config['SAMPLE_UPLOAD_FORM']
        sample_upload_file_path = os.path.join(file_path, sample_upload_file)
        xfile = os.path.join(temp, 'defect_list.xlsx')
        if (os.path.isdir(temp)):
            shutil.rmtree(temp)
        defect_list = db.session.execute(
            """
                SELECT sample.id as s_id, defect_type.name as dt_name, defect_type.primary_type as dt_pname, 
                defect_type.secondary_type as dt_sname, defect.id as d_id, defect.defect_image_name as d_image
                FROM sample
                INNER JOIN defect
                    on defect.sample = sample.id
                INNER JOIN defect_type
                    on defect.defect_type = defect_type.id
                WHERE  sample.id={}
                ;
            """.format(sample_id)
        ).fetchall()
        os.mkdir(temp)
        xfile = shutil.copyfile(sample_upload_file_path, xfile)
        wb = load_workbook(filename=xfile)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        images = []
        for defect in defect_list:
            ws.cell(row=row, column=1).value = defect.s_id
            ws.cell(row=row, column=2).value = defect.dt_name + \
                ' | ' + defect.dt_pname + ' | ' + defect.dt_sname
            ws.cell(row=row, column=3).value = '=HYPERLINK("{}", "{}")'.format(
                os.path.join(app.config['IMAGE_FOLDER'],
                             defect.d_image.split('/', 2)[-1]),
                defect.d_image.split('/', 2)[-1]
            )
            images.append(defect.d_image)
            row += 1
        wb.save(xfile)
        defect_image_folder = os.path.join(
            app.config['IMAGE_FOLDER'],
            app.config['DEFECT_IMAGE_FOLDER']
        ).replace("\\", "/")
        component_zip_file = BytesIO()
        with zipfile.ZipFile(component_zip_file, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(xfile, os.path.basename(xfile))
            for ifile in images:
                arcname = os.path.join(
                    app.config['IMAGE_FOLDER'],
                    ifile.split('/', 2)[-1]
                )
                image_path = os.path.join(defect_image_folder, ifile)
                if os.path.isfile(image_path):
                    zf.write(image_path, arcname)
        component_zip_file.seek(0)
        os.chdir(old_working_directory)
        try:
            Defect.query.filter_by(sample=sample_id).delete()
            Sample.query.filter_by(id=sample_id).delete()
            db.session.commit()
            for ifile in images:
                image_path = os.path.join(defect_image_folder, ifile)
                if os.path.isfile(image_path):
                    os.remove(image_path)
            return send_file(component_zip_file, attachment_filename='defects.zip', as_attachment=True)
        except Exception as exception:
            db.session.rollback()
            flash('Could not download the files. Try again.', 'error')
    return redirect(url_for('edit_defect_sample'))


@app.route('/ajax/getproducts', methods=('GET', 'POST'))
@approval_required
def get_products():
    if (request.method == 'POST' and 'manufacturer' in request.form):
        manufacturer_id = int(request.form.get('manufacturer'))
        products = None
        if (manufacturer_id == -1):
            products = Product.query.all()
        else:
            products = db.session.query(
                Product.id,
                Product.name
            ).filter(
                Product.manufacturer == manufacturer_id
            ).all()
        product_schema = ProductSchema(many=True)
        products_json = product_schema.dump(products).data
        if (products):
            return jsonify({'products': products_json})
    return jsonify({})


@app.route('/edit/products/display/', defaults={'product_id': None}, methods=('GET', 'POST'))
@app.route('/edit/products/display/<int:product_id>/', methods=['GET', 'POST'])
@contributor_required
def display_edit_product(product_id):
    if (product_id or (request.method == 'POST' and request.form.get('product') != '-1')):
        error = None
        product_id = product_id if (
            product_id) else request.form.get('product')
        product = Product.query.filter_by(
            id=product_id
        ).first()
        selected_manufacturer = db.session.query(
            Manufacturer.id.label('id'),
            Product.id
        ).filter(
            Manufacturer.id == Product.manufacturer
        ).filter(
            Product.id == product_id
        ).first()
        manufacturers = Manufacturer.query.all()
        form = EditProductForm()
        form.product.default = product.name
        form.total_samples.default = product.total_samples
        form.process()
        return render_template(
            'inventory/edit_product.html',
            manufacturers=manufacturers,
            selected_manufacturer=selected_manufacturer,
            product_id=product_id,
            error=error,
            form=form
        )
    return redirect(url_for('select_edit_product'))


@app.route('/edit/products/<int:product_id>/', methods=['GET', 'POST'])
@contributor_required
def edit_product(product_id):
    form = EditProductForm()
    if form.validate_on_submit():
        try:
            product = Product.query.filter_by(
                id=product_id
            ).first()
            product.manufacturer = int(request.form.get('manufacturer'))
            product.name = form.product.data
            product.total_samples = form.total_samples.data
            if request.files.getlist('pspec')[0]:
                pspec_file = request.files.getlist('pspec')[0]
                pspec_path = pspec.save(pspec_file)
                product.spec_file_name = pspec_path
            if request.files.getlist('pimage')[0]:
                pimage_file = request.files.getlist('pimage')[0]
                pimage_path = pimage.save(pimage_file)
                product.spec_image_name = pimage_path
            db.session.commit()
            return redirect(url_for('product_report', product_id=product_id))
        except Exception as exp:
            logging.error(exp)
            error = 'Error updating the Product specs. Try again.'
            flash(error, 'error')
    return redirect(url_for('display_edit_product'))


@app.route('/view/products/report/', defaults={'product_id': None}, methods=('GET', 'POST'))
@app.route('/view/products/report/<int:product_id>/', methods=['GET', 'POST'])
@approval_required
def product_report(product_id):
    error = None
    reports = []
    if (product_id or (request.method == 'POST' and request.form.get('product') != '-1')):
        product_id = product_id if (
            product_id) else request.form.get('product')

        product = Product.query.filter_by(
            id=product_id
        ).first()

        manufacturer = db.session.query(
            Manufacturer.name.label('name'),
            Product.id
        ).filter(
            Manufacturer.id == Product.manufacturer
        ).filter(
            Product.id == product_id
        ).first()

        defects_list = get_product_defects_list(product_id)

        def_rep = {}
        defect_id_set = set()
        for item in defects_list:
            dmap = def_rep.get(item.dt_id, {})
            dmap['name'] = item.dt_name
            dmap['total_samples'] = item.p_total_samples
            sample_set = dmap.get('samples', set())
            sample_set.add(item.s_sample_id)
            dmap['samples'] = sample_set
            image_list = dmap.get('images', list())
            image_list.append(item.d_image)
            dmap['images'] = image_list
            defect_id_set.add(item.dt_id)
            def_rep[item.dt_id] = dmap

        all_sample_count = get_total_samples_from_all_products()
        total_defect_sample_count = get_sample_count_for_defect(defect_id_set)

        all_product_frequency_list = []
        for dt_id in def_rep.keys():
            freq_map = {}
            freq_map['defect_type_id'] = dt_id
            freq_map['defect_name'] = def_rep[dt_id].get('name')
            freq_map['images'] = def_rep[dt_id].get('images')
            occurences = len(def_rep[dt_id].get('samples'))
            freq_map['occurences'] = occurences
            freq_map['in_frequency'] = round(
                ((float(occurences) / def_rep[dt_id].get('total_samples')) * 100), 2)
            freq_map['all_frequency'] = round(
                ((float(total_defect_sample_count[dt_id]) / float(all_sample_count)) * 100), 2)
            all_product_frequency_list.append(freq_map)

        defect_image_folder = os.path.join(
            app.config['IMAGE_FOLDER'],
            app.config['DEFECT_IMAGE_FOLDER']
        ).replace("\\", "/")

        return render_template(
            'inventory/product_report.html',
            manufacturer=manufacturer,
            product=product,
            pspec=pspec,
            pimage=pimage,
            reports=all_product_frequency_list,
            defect_image_folder=defect_image_folder,
            error=error
        )
    return redirect(url_for('select_view_product'))


@app.route('/view/defects/select/')
@approval_required
def select_defect():
    manufacturers = db.session.query(Manufacturer.id, Manufacturer.name).all()
    products = db.session.query(Product.id, Product.name).all()
    defect_types = db.session.query(DefectType.id, DefectType.name).all()
    error = None
    return render_template(
        'inventory/select_defect.html',
        manufacturers=manufacturers,
        products=products,
        defect_types=defect_types,
        error=error
    )


@app.route('/ajax/selectedmanufacturer', methods=('GET', 'POST'))
@approval_required
def selectedmanufacturer():
    if (request.method == 'POST' and 'manufacturer' in request.form):
        manufacturer_id = int(request.form.get('manufacturer'))
        products_json = None
        products = None
        defect_types = None
        defect_type_json = None
        if (manufacturer_id == -1):
            products = db.session.query(Product.id, Product.name).all()
            defect_types = db.session.query(
                DefectType.id, DefectType.name).all()
        else:
            products = db.session.query(
                Product.id,
                Product.name
            ).filter(
                Product.manufacturer == manufacturer_id
            ).all()
            defect_types = db.session.execute(
                """
                    SELECT DISTINCT(defect_type.id), defect_type.name
                    FROM defect_type
                    INNER JOIN defect
                        on defect.defect_type = defect_type.id
                    INNER JOIN chip
                        on defect.chip = chip.id
                    INNER JOIN product
                        on chip.product = product.id
                    INNER JOIN manufacturer
                        on chip.manufacturer = manufacturer.id
                    WHERE  manufacturer.id={}
                    ;
                """.format(manufacturer_id)
            ).fetchall()
        product_schema = ProductSchema(many=True)
        products_json = product_schema.dump(products).data
        defect_type_schema = DefectTypeSchema(many=True)
        defect_type_json = defect_type_schema.dump(defect_types).data
        if (products_json and defect_type_json):
            return jsonify({'products': products_json, 'defect_types': defect_type_json})
    return jsonify({})


@app.route('/ajax/selectedproduct', methods=('GET', 'POST'))
@approval_required
def selectedproduct():
    if (request.method == 'POST' and ('manufacturer', 'product' in request.form)):
        manufacturer_id = int(request.form.get('manufacturer'))
        product_id = int(request.form.get('product'))
        defect_types = None
        defect_type_json = None
        if (manufacturer_id == -1):
            defect_types = db.session.execute(
                """
                    SELECT DISTINCT(defect_type.id), defect_type.name
                    FROM defect_type
                    INNER JOIN defect
                        on defect.defect_type = defect_type.id
                    INNER JOIN chip
                        on defect.chip = chip.id
                    INNER JOIN product
                        on chip.product = product.id
                    INNER JOIN manufacturer
                        on chip.manufacturer = manufacturer.id
                    WHERE  product.id={}
                    ;
                """.format(product_id)
            ).fetchall()
        else:
            defect_types = db.session.execute(
                """
                    SELECT DISTINCT(defect_type.id), defect_type.name
                    FROM defect_type
                    INNER JOIN defect
                        on defect.defect_type = defect_type.id
                    INNER JOIN chip
                        on defect.chip = chip.id
                    INNER JOIN product
                        on chip.product = product.id
                    INNER JOIN manufacturer
                        on chip.manufacturer = manufacturer.id
                    WHERE  manufacturer.id={}
                    AND product.id={}
                    ;
                """.format(manufacturer_id, product_id)
            ).fetchall()
        defect_type_schema = DefectTypeSchema(many=True)
        defect_type_json = defect_type_schema.dump(defect_types).data
        if (defect_type_json):
            return jsonify({'defect_types': defect_type_json})
    return jsonify({})


@app.route('/view/defects/report/',  defaults={'defect_id': None}, methods=('GET', 'POST'))
@app.route('/view/defects/report/<int:defect_id>/')
@approval_required
def defect_report(defect_id):
    error = None
    if defect_id or (request.method == 'POST' and request.form.get('defect_type') != '-1'):
        defect_id = defect_id if (
            defect_id) else request.form.get('defect_type')

        defect_type = DefectType.query.filter_by(
            id=defect_id
        ).first()

        defect_types_count = get_defect_sample_occurences(defect_type.id)

        def_rep = []
        product_id_set = set()
        for item in defect_types_count:
            def_map = {}
            def_map['id'] = item.pid
            product_id_set.add(item.pid)
            def_map['name'] = item.pname
            def_map['manufacturer'] = item.mname
            def_map['occurences'] = item.occurences
            def_rep.append(def_map)

        product_total_samples_map = get_product_total_samples_map(product_id_set)

        reports = []
        for def_map in def_rep:
            def_map['frequency'] = round(
                ((float(def_map['occurences']) / product_total_samples_map[def_map['id']]) * 100), 2)
            reports.append(def_map)

        return render_template(
            'inventory/defect_report.html',
            defect_type=defect_type,
            reports=reports,
            error=error
        )

    return redirect(url_for('select_defect'))


@app.route('/view/defectslist/select/', methods=('GET', 'POST'))
@approval_required
def select_defects_list():
    manufacturers = db.session.query(Manufacturer.id, Manufacturer.name).all()
    products = db.session.query(Product.id, Product.name).all()
    defect_types = db.session.query(DefectType.id, DefectType.name).all()
    user_list = db.session.query(User.id, User.username).all()
    error = None
    return render_template(
        'inventory/select_defects_list.html',
        manufacturers=manufacturers,
        products=products,
        defect_types=defect_types,
        user_list=user_list,
        error=error
    )


@app.route('/view/defectslist/', methods=('GET', 'POST'))
# @app.route('/view/defectslist/<int:defect_id>/')
@approval_required
def defects_list():
    error = None
    if request.method == 'POST':
        m_id = int(request.form.get('manufacturer'))
        p_id = int(request.form.get('product'))
        d_id = int(request.form.get('defect_type'))
        u_id = int(request.form.get('user'))

        defects_list = get_defects_list(m_id, p_id, d_id, u_id)
        defect_image_folder = os.path.join(
            app.config['IMAGE_FOLDER'],
            app.config['DEFECT_IMAGE_FOLDER']
        ).replace("\\", "/")
        id_list = {'m_id': m_id, 'p_id': p_id, 'd_id': d_id, 'u_id': u_id}

        return render_template(
            'inventory/defects_list.html',
            defects_list=defects_list,
            id_list=id_list,
            defect_image_folder=defect_image_folder,
            error=error
        )

    return redirect(url_for('select_defect'))


@app.route('/download/images/select/', methods=('GET', 'POST'))
@approval_required
def select_image_download():
    manufacturers = db.session.query(Manufacturer.id, Manufacturer.name).all()
    products = db.session.query(Product.id, Product.name).all()
    defect_types = db.session.query(DefectType.id, DefectType.name).all()
    user_list = db.session.query(User.id, User.username).all()
    error = None
    return render_template(
        'inventory/select_image_download.html',
        manufacturers=manufacturers,
        products=products,
        defect_types=defect_types,
        user_list=user_list,
        error=error
    )


@app.route('/download/images/', methods=('GET', 'POST'))
@approval_required
def download_images():
    if request.method == 'POST':
        m_id = int(request.form.get('manufacturer'))
        p_id = int(request.form.get('product'))
        d_id = int(request.form.get('defect_type'))
        u_id = int(request.form.get('user'))
        try:
            zip_file = get_image_zip(m_id, p_id, d_id, u_id)
            return send_file(zip_file, attachment_filename='capsule.zip', as_attachment=True)
        except Exception as exception:
            flash('Error downloading images!', 'error')
    return redirect(url_for('select_image_download'))

def get_product_defects_list(product_id):
    objects = db.session.execute(
        """
            SELECT product.id as p_id, product.total_samples as p_total_samples,
            sample.sample_id as s_sample_id,
            defect_type.id as dt_id, defect_type.name as dt_name,
            defect.defect_image_name as d_image
            FROM manufacturer, product, chip, sample, defect_type, defect
            WHERE manufacturer.id = product.manufacturer
            AND product.id = chip.product
            AND chip.id = sample.chip
            AND defect_type.id = defect.defect_type
            AND defect.chip = chip.id
            AND defect.sample = sample.id
            AND product.id = {}
        ;
        """.format(product_id)
    ).fetchall()
    return objects


def get_image_zip(m_id, p_id, d_id, u_id):
    images = get_images(m_id, p_id, d_id, u_id)
    zip_file = compress_files(images)
    return zip_file


def compress_files(images):
    path = os.path.join(app.config['APP_ROOT'], 'static')
    old_working_directory = os.getcwd()
    os.chdir(path)
    defect_image_folder = os.path.join(
        app.config['IMAGE_FOLDER'],
        app.config['DEFECT_IMAGE_FOLDER']
    ).replace("\\", "/")
    component_zip_file = BytesIO()
    with zipfile.ZipFile(component_zip_file, "w", zipfile.ZIP_DEFLATED) as zf:
        for ifile in images:
            file = ifile.file
            arcname = os.path.join(
                app.config['IMAGE_FOLDER'],
                file.split('/', 2)[-1]
            )
            image_path = os.path.join(defect_image_folder, file)
            if os.path.isfile(image_path):
                zf.write(image_path, arcname)
    component_zip_file.seek(0)
    os.chdir(old_working_directory)
    return component_zip_file


def get_images(m_id, p_id, d_id, u_id):
    images = None
    if (u_id == -1):
        images = get_images_for_default_user(m_id, p_id, d_id)
    elif (m_id == -1 and p_id == -1 and d_id == -1):
        images = query_images_by_all_for_user(u_id)
    elif (m_id > -1 and p_id > -1 and d_id > -1):
        images = query_images_by_three_filters_for_user(m_id, p_id, d_id, u_id)
    elif (m_id > -1 and p_id > -1):
        images = query_images_by_m_p_for_user(m_id, p_id, u_id)
    elif (m_id > -1 and d_id > -1):
        images = query_images_by_m_d_for_user(m_id, d_id, u_id)
    elif (p_id > -1 and d_id > -1):
        images = query_images_by_p_d_for_user(p_id, d_id, u_id)
    elif (m_id > -1):
        images = query_images_by_m_for_user(m_id, u_id)
    elif (p_id > -1):
        images = query_images_by_p_for_user(p_id, u_id)
    elif (d_id > -1):
        images = query_images_by_d_for_user(d_id, u_id)
    return images


def get_images_for_default_user(m_id, p_id, d_id):
    images = None
    if (m_id == -1 and p_id == -1 and d_id == -1):
        images = query_images_by_all()
    elif (m_id > -1 and p_id > -1 and d_id > -1):
        images = query_images_by_three_filters(m_id, p_id, d_id)
    elif (m_id > -1 and p_id > -1):
        images = query_images_by_m_p(m_id, p_id)
    elif (m_id > -1 and d_id > -1):
        images = query_images_by_m_d(m_id, d_id)
    elif (p_id > -1 and d_id > -1):
        images = query_images_by_p_d(p_id, d_id)
    elif (m_id > -1):
        images = query_images_by_m(m_id)
    elif (p_id > -1):
        images = query_images_by_p(p_id)
    elif (d_id > -1):
        images = query_images_by_d(d_id)
    return images


def query_images_by_all():
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            ;
        """
    ).fetchall()
    return objects


def get_total_samples_from_all_products():
    objects = db.session.query(
            func.sum(Product.total_samples).label('total_samples')
        ).first()
    return objects[0]


def get_sample_count_for_defect(defect_ids):
    defect_ids_list = '(' + ",".join([str(i) for i in defect_ids]) + ')'
    objects = db.session.execute(
        """
        SELECT defect_type.id as id, 
        count(DISTINCT sample.sample_id) as total_samples
        FROM sample, defect, defect_type
        WHERE defect.defect_type = defect_type.id
        AND defect.sample = sample.id
        AND defect_type.id IN {}
        GROUP BY defect_type.id;
        """.format(defect_ids_list)
    ).fetchall()
    sample_count_for_defect_map = {
            f.id: f.total_samples
            for f in objects
    }
    return sample_count_for_defect_map


def get_defect_sample_occurences(defect_type_id):
    objects = db.session.execute(
        """
        SELECT manufacturer.name as mname, 
        product.name as pname, product.id as pid,  
        count(DISTINCT sample.sample_id) as occurences
        FROM defect_type, defect, sample, chip, product, manufacturer
        WHERE defect_type.id = defect.defect_type
        AND defect.sample = sample.id
        And defect.chip = chip.id
        And chip.manufacturer = manufacturer.id
        And chip.product = product.id
        AND defect_type.id = {}
        GROUP BY product.id;
        """.format(defect_type_id)
    ).fetchall()
    return objects

def get_product_total_samples_map(product_ids):
    product_ids_list = '(' + ",".join([str(i) for i in product_ids]) + ')'
    objects = db.session.execute(
        """
        SELECT product.id, product.total_samples
        FROM product
        WHERE product.id IN {};
        """.format(product_ids_list)
    ).fetchall()
    product_total_samples_map = {
            f.id: f.total_samples
            for f in objects
    }
    return product_total_samples_map


def query_images_by_all_for_user(u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE chip.user= {}
            ;
        """.format(u_id)
    ).fetchall()
    return objects


def query_images_by_three_filters(m_id, p_id, d_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND product.id={}
            AND defect_type.id={}
            ;
        """.format(m_id, p_id, d_id)
    ).fetchall()
    return objects


def query_images_by_three_filters_for_user(m_id, p_id, d_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND product.id={}
            AND defect_type.id={}
            AND chip.user= {}
            ;
        """.format(m_id, p_id, d_id, u_id)
    ).fetchall()
    return objects


def query_images_by_m_p(m_id, p_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND product.id={}
            ;
        """.format(m_id, p_id)
    ).fetchall()
    return objects


def query_images_by_m_p_for_user(m_id, p_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND product.id={}
            AND chip.user= {}
            ;
        """.format(m_id, p_id, u_id)
    ).fetchall()
    return objects


def query_images_by_m_d(m_id, d_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND defect_type.id={}
            ;
        """.format(m_id, d_id)
    ).fetchall()
    return objects


def query_images_by_m_d_for_user(m_id, d_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND defect_type.id={}
            AND chip.user= {}
            ;
        """.format(m_id, d_id, u_id)
    ).fetchall()
    return objects


def query_images_by_p_d_for_user(p_id, d_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            WHERE product.id={}
            AND defect_type.id={}
            ;
        """.format(p_id, d_id, u_id)
    ).fetchall()
    return objects


def query_images_by_p_d(p_id, d_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            WHERE product.id={}
            AND defect_type.id={}
            ;
        """.format(p_id, d_id)
    ).fetchall()
    return objects


def query_images_by_m(m_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            ;
        """.format(m_id)
    ).fetchall()
    return objects


def query_images_by_m_for_user(m_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id={}
            AND chip.user= {}
            ;
        """.format(m_id, u_id)
    ).fetchall()
    return objects


def query_images_by_p(p_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            WHERE product.id={}
            ;
        """.format(p_id)
    ).fetchall()
    return objects


def query_images_by_p_for_user(p_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            WHERE product.id={}
            AND chip.user= {}
            ;
        """.format(p_id, u_id)
    ).fetchall()
    return objects


def query_images_by_d(d_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            WHERE defect_type.id={}
            ;
        """.format(d_id)
    ).fetchall()
    return objects


def query_images_by_d_for_user(d_id, u_id):
    objects = db.session.execute(
        """
            SELECT defect.defect_image_name as file
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            WHERE defect_type.id={}
            AND chip.user= {}
            ;
        """.format(d_id, u_id)
    ).fetchall()
    return objects


def get_defects_list(m_id, p_id, d_id, u_id):
    defects_list = None
    if (u_id == -1):
        defects_list = get_defects_list_for_default_user(m_id, p_id, d_id)
    elif (m_id == -1 and p_id == -1 and d_id == -1):
        defects_list = query_defects_by_all_for_user(u_id)
    elif (m_id > -1 and p_id > -1 and d_id > -1):
        defects_list = query_defects_by_three_filters_for_user(
            m_id, p_id, d_id, u_id)
    elif (m_id > -1 and p_id > -1):
        defects_list = query_defects_by_m_p_for_user(m_id, p_id, u_id)
    elif (m_id > -1 and d_id > -1):
        defects_list = query_defects_by_m_d_for_user(m_id, d_id, u_id)
    elif (p_id > -1 and d_id > -1):
        defects_list = query_defects_by_p_d_for_user(p_id, d_id, u_id)
    elif (m_id > -1):
        defects_list = query_defects_by_m_for_user(m_id, u_id)
    elif (p_id > -1):
        defects_list = query_defects_by_p_for_user(p_id, u_id)
    elif (d_id > -1):
        defects_list = query_defects_by_d_for_user(d_id, u_id)
    return defects_list


def get_defects_list_for_default_user(m_id, p_id, d_id):
    defects_list = None
    if (m_id == -1 and p_id == -1 and d_id == -1):
        defects_list = query_defects_by_all()
    elif (m_id > -1 and p_id > -1 and d_id > -1):
        defects_list = query_defects_by_three_filters(m_id, p_id, d_id)
    elif (m_id > -1 and p_id > -1):
        defects_list = query_defects_by_m_p(m_id, p_id)
    elif (m_id > -1 and d_id > -1):
        defects_list = query_defects_by_m_d(m_id, d_id)
    elif (p_id > -1 and d_id > -1):
        defects_list = query_defects_by_p_d(p_id, d_id)
    elif (m_id > -1):
        defects_list = query_defects_by_m(m_id)
    elif (p_id > -1):
        defects_list = query_defects_by_p(p_id)
    elif (d_id > -1):
        defects_list = query_defects_by_d(d_id)
    return defects_list


def query_defects_by_all():
    objects = db.session.execute(
        """
            SELECT manufacturer.id as m_id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            ;
        """
    ).fetchall()
    return objects


def query_defects_by_all_for_user(u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id as m_id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  chip.user= {}
            ;
        """.format(u_id)
    ).fetchall()
    return objects


def query_defects_by_three_filters(m_id, p_id, d_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND product.id={}
            AND defect_type.id= {}
            ;
        """.format(m_id, p_id, d_id)
    ).fetchall()
    return objects


def query_defects_by_three_filters_for_user(m_id, p_id, d_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND product.id={}
            AND defect_type.id= {}
            AND chip.user= {}
            ;
        """.format(m_id, p_id, d_id, u_id)
    ).fetchall()
    return objects


def query_defects_by_m_p(m_id, p_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND product.id={}
            ;
        """.format(m_id, p_id)
    ).fetchall()
    return objects


def query_defects_by_m_p_for_user(m_id, p_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND product.id={}
            AND chip.user= {}
            ;
        """.format(m_id, p_id, u_id)
    ).fetchall()
    return objects


def query_defects_by_m_d(m_id, d_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND defect_type.id= {}
            ;
        """.format(m_id, d_id)
    ).fetchall()
    return objects


def query_defects_by_m_d_for_user(m_id, d_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND defect_type.id= {}
            AND chip.user= {}
            ;
        """.format(m_id, d_id, u_id)
    ).fetchall()
    return objects


def query_defects_by_p_d(p_id, d_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE product.id={}
            AND defect_type.id= {}
            ;
        """.format(p_id, d_id)
    ).fetchall()
    return objects


def query_defects_by_p_d_for_user(p_id, d_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE product.id={}
            AND defect_type.id= {}
            AND chip.user= {}
            ;
        """.format(p_id, d_id, u_id)
    ).fetchall()
    return objects


def query_defects_by_m(m_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            ;
        """.format(m_id)
    ).fetchall()
    return objects


def query_defects_by_m_for_user(m_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE  manufacturer.id= {}
            AND chip.user= {}
            ;
        """.format(m_id, u_id)
    ).fetchall()
    return objects


def query_defects_by_p(p_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE product.id={}
            ;
        """.format(p_id)
    ).fetchall()
    return objects


def query_defects_by_p_for_user(p_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE product.id={}
            AND chip.user= {}
            ;
        """.format(p_id, u_id)
    ).fetchall()
    return objects


def query_defects_by_d(d_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE defect_type.id= {}
            ;
        """.format(d_id)
    ).fetchall()
    return objects


def query_defects_by_d_for_user(d_id, u_id):
    objects = db.session.execute(
        """
            SELECT manufacturer.id, manufacturer.name as m_name,
            product.id as p_id, product.name as p_name,
            defect_type.id as d_id,
            defect_type.name as d_name,
            defect_type.primary_type as d_primary,
            defect_type.secondary_type as d_secondary,
            defect.defect_image_name as d_image
            FROM defect
            INNER JOIN defect_type
                on defect.defect_type = defect_type.id
            INNER JOIN chip
                on defect.chip = chip.id
            INNER JOIN product
                on chip.product = product.id
            INNER JOIN manufacturer
                on chip.manufacturer = manufacturer.id
            WHERE defect_type.id= {}
            AND chip.user= {}
            ;
        """.format(d_id, u_id)
    ).fetchall()
    return objects
