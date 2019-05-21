from counterfeit_ic import app, db, pimage, pspec, archives
from flask import render_template, redirect, session, request, url_for, flash, abort
from openpyxl import load_workbook
from user.models import User
from inventory.models import Manufacturer, Product, DefectType, Chip, Sample, Defect
from inventory.forms import CreateComponentsForm, AddComponentsForm
from flask_uploads import UploadSet, IMAGES, configure_uploads
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from user.decorators import contributor_required, approval_required

import pyexcel as excel
import zipfile
import os
import shutil
import glob
import datetime
import pathlib
import logging


@app.route('/create/manufacturer', methods=('GET', 'POST'))
@contributor_required
def create_manufacturer():
    return render_template('inventory/create_manufacturer.html')


@app.route('/create/components', methods=('GET', 'POST'))
@contributor_required
def create_components():
    form = CreateComponentsForm()
    if form.validate_on_submit() and ('pspec', 'pimage', 'archive' in request.files):
        manufacturer = process_manufacturer(form)
        product = process_product(form, request, manufacturer)
        chip = get_or_create_model(Chip, user=session.get(
            'id'), manufacturer=manufacturer.id, product=product.id)
        processed_count, error_count = process_archive(
            request, manufacturer, product, chip)
        if error_count > 0:
            error = '{} files successfully added. {} files ignored.'.format(
                processed_count, error_count)
            flash(error, 'error')
        else:
            success = '{} files successfully added.'.format(processed_count)
            flash(success, 'success')
    return render_template('inventory/create_components.html', form=form)


def process_manufacturer(form):
    return get_or_create_model(
        Manufacturer,
        name=form.manufacturer.data
    )


def process_product(form, request, manufacturer):
    pspec_file = request.files.getlist('pspec')[0]
    pimage_file = request.files.getlist('pimage')[0]
    logging.info(pspec_file)
    logging.info(pimage_file)
    pspec_path = pspec.save(pspec_file)
    logging.info(pspec_path)
    pimage_path = pimage.save(pimage_file)
    logging.info(pimage_path)
    product = Product(
        name=form.product.data,
        manufacturer_id=manufacturer.id,
        spec_file_name=pspec_path,
        spec_image_name=pimage_path,
    )
    db.session.add(product)
    db.session.flush()
    return product


def get_or_create_model(model, **kwargs):
    instance = db.session.query(model).filter_by(**kwargs).first()
    if not instance:
        instance = model(**kwargs)
        db.session.add(instance)
        db.session.flush()
    return instance


@app.route('/create/add_files', methods=('GET', 'POST'))
@contributor_required
def add_files():
    form = CreateComponentsForm()
    if form.validate_on_submit():  # and 'photo' in request.files:
        if('photo', 'pdf' in request.files):
            pname = request.files.getlist('photo')[0]
            pname = pimage.save(pname)
            dname = request.files.getlist('pdf')[0]
            dname = pspec.save(dname)
            pass
    else:
        flash('Files could not be added!', 'error')
    return render_template('inventory/create_product.html', form=form)


def process_archive(request, manufacturer, product, chip):
    temp = app.config['TEMP']
    image_folder = app.config['IMAGE_FOLDER']
    defect_image_folder = app.config['DEFECT_IMAGE_FOLDER']
    destination = datetime.datetime.now().strftime('%Y%m%d')
    row_start = app.config['ROW_START']
    path = os.path.join(app.config['APP_ROOT'], 'static')
    delim = '_'
    archive = request.files.getlist('archive')[0]

    old_working_directory = os.getcwd()
    os.chdir(path)

    if (os.path.isdir(temp)):
        shutil.rmtree(temp)

    index = None
    zfile = zipfile.ZipFile(archive, "r")
    zfile.extractall(os.path.join(temp))
    if (not os.path.isdir(os.path.join(image_folder, defect_image_folder, destination))):
        index = 0
        pathlib.Path(
            os.path.join(image_folder, defect_image_folder, destination)
        ).mkdir(parents=True, exist_ok=True)
    else:
        index = len(os.listdir(os.path.join(
            image_folder, defect_image_folder, destination)))

    xfiles = glob.glob(os.path.join(temp, '*.xlsx'))
    xfile = xfiles[0] if (xfiles) else None
    wb = load_workbook(filename=xfile)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])

    error_count = 0
    processed_count = 0
    for row in range(row_start, ws.max_row+1):
        try:
            sample_id = int(ws.cell(row=row, column=1).value)
            defect = ws.cell(row=row, column=2).value
            image = ws.cell(row=row, column=3).value
            defect_name = defect.split('|', 2)[0].strip()
            primary_type = defect.split('|', 2)[1].strip()
            secondary_type = defect.split('|', 2)[2].strip()
            image_path = os.path.join(temp, image).replace("\\", "/")
            img_ext = os.path.splitext(image_path)[1].lower()
            if (not os.path.isfile(image_path)):
                logging.info("Image path doesn't exist: " + image_path)
                continue

            manufacturer_name = manufacturer.name if (
                len(manufacturer.name) <= 10) else manufacturer.name[0:10]
            product_name = product.name if (
                len(product.name) <= 10) else product.name[0:10]
            secure_image = secure_filename(
                manufacturer_name + delim + product_name + delim +
                defect_name.lower() + delim + str(index) + img_ext
            )
            dest = os.path.join(
                destination,
                secure_image
            ).replace("\\", "/")

            os.rename(image_path, os.path.join(
                image_folder, defect_image_folder, dest))
            index += 1
            defect_type = get_or_create_model(
                DefectType,
                name=defect_name,
                primary_type=primary_type,
                secondary_type=secondary_type
            )
            sample = get_or_create_model(
                Sample,
                sample_id=sample_id,
                chip=chip.id
            )
            defect = Defect(
                chip_id=chip.id,
                sample_id=sample.id,
                defect_type_id=defect_type.id,
                defect_image_name=dest
            )
            db.session.add(defect)
            processed_count += 1
        except Exception as exception:
            logging.error(exception)
            error_count += 1
            return (processed_count, error_count)
    if (processed_count > 0):
        db.session.commit()
    os.chdir(old_working_directory)
    return (processed_count, error_count)


@app.route('/add/components', methods=('GET', 'POST'))
@contributor_required
def add_components():
    error = None
    notification = None
    form = AddComponentsForm()
    manufacturers = Manufacturer.query.all()
    if (form.validate_on_submit() and ('archive' in request.files)):
        manufacturer_id = int(request.form.get('manufacturer'))
        product_id = int(request.form.get('product'))
        if (manufacturer_id == -1 or product_id == -1):
            error = 'Manufacturer or Product not selected'
            flash(error, 'error')
        else:
            manufacturer = Manufacturer.query.filter_by(
                id=manufacturer_id
            ).first()
            product = Product.query.filter_by(
                id=product_id
            ).first()
            chip = get_or_create_model(
                Chip, user=session.get('id'),
                manufacturer=manufacturer.id, product=product.id
            )
            processed_count, error_count = process_archive(
                request, manufacturer, product, chip)
            if error_count > 0:
                error = '{} files successfully added. {} files ignored.'.format(
                    processed_count, error_count)
                flash(error, 'error')
            else:
                success = '{} files successfully added.'.format(
                    processed_count)
                flash(success, 'success')
    return render_template(
        'inventory/add_components.html',
        manufacturers=manufacturers,
        form=form,
    )
